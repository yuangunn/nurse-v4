from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import json
import queue
import logging
import traceback
import threading

from . import solver_progress  # solver-agnostic progress/cancel registry

logger = logging.getLogger(__name__)

from . import database as db
from . import profiles as prof
from .models import GenerateRequest, ScheduleSave, ScoringRule, Nurse, Rules
from .scheduler import NurseScheduler

# ── 현재 활성 프로필 ──────────────────────────────────────────────────────────
_current_profile_id: Optional[str] = None
_current_profile_password: Optional[str] = None

# ── HiGHS 인스턴스 추적 (중지 기능용) ────────────────────────────────────────
# PuLP의 HiGHS 솔버가 내부적으로 생성하는 highspy.Highs 인스턴스를 가로채
# cancelSolve() 호출과 mip_gap 캡처를 가능하게 함.
_solve_lock = threading.Lock()
_last_mip_gap: Optional[float] = None
_log_queue: queue.Queue = queue.Queue()
_last_generate_result: Optional[dict] = None  # 마지막 생성 결과 보관 (새로고침 복구용)

try:
    import highspy as _highspy_mod
    _OrigHighs = _highspy_mod.Highs

    class _HighsAdapter:
        """solver_progress 어댑터 — HiGHS 라이브 인스턴스 래핑 (cancel/progress)."""
        def __init__(self, highs):
            self._h = highs

        def cancel(self):
            try:
                self._h.cancelSolve()
            except Exception:
                pass

        def progress(self) -> dict:
            import math
            out = {"gap_percent": None, "nodes": 0, "has_solution": False, "is_running": True}
            try:
                status, gap = self._h.getInfoValue("mip_gap")
                if status.value == 0 and math.isfinite(float(gap)):
                    out["gap_percent"] = round(float(gap) * 100, 2)
                    out["has_solution"] = True
                status2, nodes = self._h.getInfoValue("mip_node_count")
                if status2.value == 0:
                    out["nodes"] = int(nodes)
            except Exception:
                pass
            return out

    # 솔버 로그 패턴 → 사용자 행동 힌트 (run당 키별 1회만)
    import re as _re
    _LOG_HINTS = [
        (_re.compile(r"Presolve\s*:?.*[Ii]nfeasible"), "presolve_infeasible",
         "💡 전처리에서 모순 감지 — 사전입력끼리 충돌일 가능성이 높습니다. 실패 시 '정밀 충돌 분석'을 실행해 보세요."),
        (_re.compile(r"kTimeLimit|[Tt]ime limit reached"), "time_limit",
         "💡 시간 제한 도달 — 현재 해는 표시된 gap% 이내의 품질입니다."),
        (_re.compile(r"kInterrupt"), "interrupted",
         "💡 중지 신호로 탐색이 종료되었습니다."),
    ]

    class _TrackableHighs(_OrigHighs):
        def run(self):
            global _last_mip_gap, _log_queue
            # 큐 초기화 (이전 실행 잔여 로그 제거)
            while not _log_queue.empty():
                try: _log_queue.get_nowait()
                except Exception: break
            # 로그 콜백 등록 — 솔버 출력을 큐에 적재
            # highspy 1.8+: cbLogging.subscribe(fn) — event.message로 로그 수신
            _hinted = set()

            def _on_log(event):
                msg = getattr(event, "message", "")
                if msg and msg.strip():
                    _log_queue.put({"type": "log", "msg": msg.rstrip()})
                    for rx, key, hint in _LOG_HINTS:
                        if key not in _hinted and rx.search(msg):
                            _hinted.add(key)
                            _log_queue.put({"type": "hint", "msg": hint})
            try:
                self.cbLogging.subscribe(_on_log)
            except Exception:
                pass
            self.setOptionValue("output_flag", True)
            # HandleUserInterrupt=True 필수: 이 플래그가 있어야 cancelSolve()가
            # MIP 인터럽트 콜백을 활성화하여 솔버를 실제로 중단할 수 있음
            self.HandleUserInterrupt = True
            _hadapter = _HighsAdapter(self)
            solver_progress.register(_hadapter)
            try:
                result = super().run()
                return result
            finally:
                try:
                    status, gap = super().getInfoValue("mip_gap")
                    if status.value == 0:  # kOk
                        _last_mip_gap = float(gap)
                except Exception:
                    pass
                solver_progress.unregister(_hadapter)

    _highspy_mod.Highs = _TrackableHighs
except ImportError:
    pass  # highspy 없으면 패스 (cancel/gap 기능 비활성화)

app = FastAPI(title="NurseScheduler v3")

# 정적 파일 서빙 (frontend/ 하위 css, js, lib)
_frontend_dir = Path(__file__).resolve().parent.parent / "frontend"
for _subdir in ("css", "js", "lib", "fonts", "assets"):
    _sub_path = _frontend_dir / _subdir
    if _sub_path.exists():
        app.mount(f"/{_subdir}", StaticFiles(directory=str(_sub_path)), name=_subdir)


# 개발 중 stylesheet/script 캐시로 인한 stale UI 방지 — 정적 파일에 no-cache
@app.middleware("http")
async def _no_cache_static(request, call_next):
    response = await call_next(request)
    path = request.url.path
    if path.startswith(("/css/", "/js/", "/lib/", "/fonts/", "/assets/")) or path == "/":
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        # ETag / Last-Modified 제거 — 브라우저 304 conditional 응답 차단
        for h in ("etag", "last-modified"):
            if h in response.headers:
                del response.headers[h]
    return response


@app.on_event("startup")
def startup():
    prof.init_default_profiles()
    # 기본 DB 초기화 (프로필 전환 전 폴백)
    db.init_db()
    # 직전 비정상 종료로 남은 게스트 임시 DB 정리
    try:
        guest_db = prof._db_path_for_profile("_guest_temp")
        if guest_db.exists():
            guest_db.unlink()
    except Exception:
        pass


@app.on_event("shutdown")
def shutdown():
    """서버 종료 시 열린 프로필을 닫아 재암호화 보장 (graceful 종료 경로)."""
    global _current_profile_id, _current_profile_password
    if _current_profile_id:
        try:
            prof.close_profile(_current_profile_id, _current_profile_password or "")
        except Exception:
            pass
        _current_profile_id = None
        _current_profile_password = None


# ── 프로필 API ────────────────────────────────────────────────────────────────

@app.get("/api/profiles")
def get_profiles():
    return {
        "profiles": prof.list_profiles(),
        "has_master_password": prof.has_master_password(),
        "current_profile": _current_profile_id,
    }


@app.post("/api/profiles/create")
def create_profile(body: dict):
    profile_id = body.get("id", "").strip()
    name = body.get("name", "").strip()
    password = body.get("password", "")
    is_guest = body.get("is_guest", False)
    if not profile_id or not name:
        raise HTTPException(400, "프로필 ID와 이름을 입력해주세요.")
    result = prof.create_profile(profile_id, name, password, is_guest)
    if not result.get("ok"):
        raise HTTPException(400, result.get("error", "프로필 생성 실패"))
    return result


@app.post("/api/profiles/open")
def open_profile(body: dict):
    global _current_profile_id, _current_profile_password
    profile_id = body.get("id", "")
    password = body.get("password", "")

    # 마스터 비밀번호 확인
    if prof.has_master_password():
        master_pw = body.get("master_password", "")
        if not master_pw:
            return {"ok": False, "need_master_password": True,
                    "error": "마스터 비밀번호를 입력해주세요."}
        if not prof.verify_master_password(master_pw):
            return {"ok": False, "error": "마스터 비밀번호가 틀렸습니다."}

    # 같은 프로필 재오픈이면 먼저 닫아 .db.enc를 최신화한 뒤 다시 연다.
    if _current_profile_id and _current_profile_id == profile_id:
        prof.close_profile(_current_profile_id, _current_profile_password or "")
        _current_profile_id = None
        _current_profile_password = None

    # 대상 프로필을 먼저 검증·복호화하고, 성공했을 때에만 현재 프로필을 닫는다.
    # (실패 시 현재 프로필을 닫아버리면 평문이 사라진 상태로 전역이 옛 경로를
    #  가리켜, 이후 요청이 빈 스텁 DB를 만들고 close가 그것을 재암호화한다.)
    result = prof.open_profile(profile_id, password)
    if not result.get("ok"):
        return result

    if _current_profile_id:
        prof.close_profile(_current_profile_id, _current_profile_password or "")

    # DB 경로 전환
    db_path = result["db_path"]
    db.get_db_path = lambda: db_path
    db.init_db()

    # 한 번 열릴 때 유령 간호사 참조 정리 (과거 데이터 호환)
    try:
        removed = db.cleanup_orphan_nurse_refs()
        if removed:
            print(f"[cleanup] {removed} orphan nurse references removed from saved data")
    except Exception as e:
        print(f"[cleanup] failed: {e}")

    _current_profile_id = profile_id
    _current_profile_password = password if not result.get("is_guest") else ""

    return {"ok": True, "profile_id": profile_id,
            "is_guest": result.get("is_guest", False)}


@app.post("/api/profiles/close")
def close_profile():
    global _current_profile_id, _current_profile_password
    if _current_profile_id:
        prof.close_profile(_current_profile_id, _current_profile_password or "")
        _current_profile_id = None
        _current_profile_password = None
    return {"ok": True}


@app.delete("/api/profiles/{profile_id}")
def delete_profile(profile_id: str):
    global _current_profile_id, _current_profile_password
    # 현재 열린 프로필이면 먼저 close 상태로 전환 (삭제 중 암호화 시도 방지)
    if _current_profile_id == profile_id:
        _current_profile_id = None
        _current_profile_password = None
    result = prof.delete_profile(profile_id)
    if not result.get("ok"):
        raise HTTPException(400, result.get("error"))
    return result


@app.post("/api/profiles/change-password")
def change_profile_password(body: dict):
    profile_id = body.get("id", "")
    old_password = body.get("old_password", "")
    new_password = body.get("new_password", "")
    force_reset = body.get("force_reset", False)
    if force_reset:
        # 개발자 모드: 비밀번호 강제 초기화 (제거)
        # 마스터 비밀번호가 설정돼 있으면 검증 — 무인증 강제 초기화 방지
        if prof.has_master_password():
            if not prof.verify_master_password(body.get("master_password", "")):
                raise HTTPException(403, "마스터 비밀번호가 필요합니다.")
        result = prof.force_reset_password(profile_id)
        if not result.get("ok"):
            raise HTTPException(400, result.get("error"))
        return result
    if not new_password:
        raise HTTPException(400, "새 비밀번호를 입력해주세요.")
    result = prof.change_password(profile_id, old_password, new_password,
                                  is_open=(_current_profile_id == profile_id))
    if not result.get("ok"):
        raise HTTPException(400, result.get("error"))
    # 현재 열린 프로필이면 비밀번호 업데이트
    global _current_profile_password
    if _current_profile_id == profile_id:
        _current_profile_password = new_password
    return result


@app.post("/api/profiles/master-password")
def set_master_password(body: dict):
    action = body.get("action", "set")
    if action == "set":
        password = body.get("password", "")
        if not password:
            raise HTTPException(400, "비밀번호를 입력해주세요.")
        prof.set_master_password(password)
        return {"ok": True}
    elif action == "remove":
        current = body.get("current_password", "")
        if prof.has_master_password() and not prof.verify_master_password(current):
            raise HTTPException(400, "현재 마스터 비밀번호가 틀렸습니다.")
        prof.remove_master_password()
        return {"ok": True}
    elif action == "verify":
        password = body.get("password", "")
        return {"ok": prof.verify_master_password(password)}
    raise HTTPException(400, "알 수 없는 action")


# ── 개발자 API ────────────────────────────────────────────────────────────────

@app.get("/api/dev/cpsat-selftest")
def dev_cpsat_selftest():
    """CP-SAT(ortools)가 번들 환경에서 임포트·풀이되는지 셀프테스트.
    Phase 0 오프라인 번들 게이트 검증용 — 번들된 exe에서 호출해 ortools 네이티브
    libs가 올바로 실렸는지 확인한다. (ortools는 런타임 네트워크를 쓰지 않음.)"""
    try:
        from ortools.sat.python import cp_model
        m = cp_model.CpModel()
        xs = [m.NewBoolVar(f"x{i}") for i in range(5)]
        m.Add(sum(xs) == 2)
        s = cp_model.CpSolver()
        st = s.Solve(m)
        ok = st in (cp_model.OPTIMAL, cp_model.FEASIBLE)
        return {
            "ok": ok,
            "sum": sum(int(s.Value(x)) for x in xs) if ok else None,
            "ortools_status": int(st),
        }
    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}


@app.get("/api/dev/info")
def dev_info():
    """현재 DB 경로, 크기, 간호사 수"""
    import os
    db_path = db.get_db_path()
    size_bytes = os.path.getsize(db_path) if os.path.exists(db_path) else 0
    if size_bytes < 1024:
        size_str = f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        size_str = f"{size_bytes / 1024:.1f} KB"
    else:
        size_str = f"{size_bytes / 1024 / 1024:.1f} MB"
    nurses = db.get_nurses()
    return {"path": db_path, "size": size_str, "nurses": len(nurses)}


@app.post("/api/dev/reset-seed")
def dev_reset_seed():
    """예시 데이터(18명) 재생성"""
    from .database import _seed_nurses
    # 삭제와 시드를 같은 트랜잭션에서 — 분리하면 시드 INSERT가 커밋되지 않고
    # GC 롤백되어 빈 테이블만 남는다.
    with db.get_conn() as conn:
        conn.execute("DELETE FROM nurses")
        _seed_nurses(conn)
    return {"ok": True}


@app.get("/api/fairness_ledger")
def get_fairness_ledger(year: int, month: int, months_back: int = 3):
    """공정성 원장 — 직전 N개월 저장본 누적 (야간/주말/공휴일 근무 + 위시)."""
    try:
        return {"ledger": db.compute_fairness_ledger(year, month, months_back),
                "wish": db.compute_wish_ledger(year, month, months_back)}
    except Exception:
        return {"ledger": {}, "wish": {}}


@app.get("/api/relax_ledger")
def get_relax_ledger(year: int, month: int, months_back: int = 3):
    """완화 이력 원장 — 직전 N개월 저장본에서 간호사별 뒤집힌 원티드 수 (M6 P2①)."""
    try:
        return db.compute_relax_ledger(year, month, months_back)
    except Exception:
        return {}


@app.get("/api/prev_month_nights")
def get_prev_month_nights(year: int, month: int):
    """직전 달 최신 저장 근무표 기준 간호사별 야간 수 — '전월N' 자동 채움용."""
    try:
        return db.compute_prev_month_nights(year, month)
    except Exception:
        return {}


@app.get("/api/generation_runs")
def get_generation_runs(limit: int = 20):
    """생성 이력 (현재 프로필 DB 기준, 최신순)."""
    try:
        return db.list_generation_runs(limit=limit)
    except Exception:
        return []


@app.get("/api/dev/download-db")
def dev_download_db():
    """현재 DB 파일 다운로드"""
    from fastapi.responses import FileResponse as FR
    db_path = db.get_db_path()
    return FR(db_path, media_type="application/octet-stream",
              filename="nurse_backup.db")


# ── 간호사 CSV 템플릿/임포트 ──────────────────────────────────────────────

_NURSE_CSV_HEADER = [
    "id", "이름", "그룹", "성별", "가능근무",
    "야간전담", "시니어리티", "주휴요일", "주휴로테이션",
    "트레이닝", "트레이닝종료일", "프리셉터ID",
    "전입일", "전출일",
]
_NURSE_CSV_EXAMPLE = [
    ["n001", "김지현", "A", "female", "DC,D,EC,E,NC,N", "N", "0", "", "Y", "N", "", "", "", ""],
    ["n002", "이수진", "A", "female", "DC,D,EC,E,NC,N", "N", "1", "목", "Y", "N", "", "", "2026-04-01", ""],
    ["n003", "박민지", "B", "male", "D,E,N", "Y", "5", "", "N", "N", "", "", "", "2026-06-30"],
    ["n004", "신입간호사", "C", "female", "D,E,N", "N", "20", "", "Y", "Y", "2026-04-30", "n001", "2026-04-01", ""],
]
_NURSE_CSV_GUIDE = [
    ["# 작성 방법:"],
    ["# id — 고유 ID (영문/숫자, 중복 불가)"],
    ["# 이름 — 간호사 이름"],
    ["# 그룹 — A/B/C 등 자유 입력"],
    ["# 성별 — female 또는 male"],
    ["# 가능근무 — 쉼표로 구분 (예: DC,D,EC,E,NC,N)"],
    ["# 야간전담 — Y(야간전담) 또는 N(일반)"],
    ["# 시니어리티 — 숫자 (작을수록 선임). 앱 내 간호사 목록 순서로 자동 결정됨 (생략 가능)"],
    ["# 주휴요일 — 일/월/화/수/목/금/토 중 하나, 또는 빈칸(임의)"],
    ["# 주휴로테이션 — Y(4주마다 당김) 또는 N(고정)"],
    ["# 트레이닝 — Y(신규) 또는 N"],
    ["# 트레이닝종료일 — YYYY-MM-DD (트레이닝=Y일 때)"],
    ["# 프리셉터ID — 트레이닝=Y일 때 담당 프리셉터의 id"],
    ["# 전입일 — YYYY-MM-DD (이 날부터 근무 가능, 빈칸=상시 근무)"],
    ["# 전출일 — YYYY-MM-DD (이 날까지 근무 가능, 빈칸=상시 근무)"],
    ["#"],
    ["# 주의: #으로 시작하는 행은 무시됩니다. 헤더 행과 데이터 행만 남기고 사용하세요."],
    ["#"],
]


@app.get("/api/nurses/template")
def nurse_template():
    """간호사 일괄 등록용 CSV 템플릿 다운로드"""
    import io
    import csv
    from fastapi.responses import Response

    buf = io.StringIO()
    # UTF-8 BOM (한글 엑셀 호환)
    buf.write("\ufeff")
    writer = csv.writer(buf)
    for row in _NURSE_CSV_GUIDE:
        writer.writerow(row)
    writer.writerow(_NURSE_CSV_HEADER)
    for row in _NURSE_CSV_EXAMPLE:
        writer.writerow(row)

    content = buf.getvalue().encode("utf-8")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="nurses_template.csv"'
        },
    )


_JUHU_REV = {0: "일", 1: "월", 2: "화", 3: "수", 4: "목", 5: "금", 6: "토"}


def _nurse_to_row(n: dict) -> list:
    """간호사 dict → 템플릿 컬럼 순서의 행 (CSV·xlsx 내보내기 공용)."""
    capable = n.get("capable_shifts", [])
    if isinstance(capable, str):
        capable = [capable]
    juhu_day = n.get("juhu_day")
    juhu_ko = _JUHU_REV.get(juhu_day, "") if juhu_day is not None else ""
    return [
        n.get("id", ""),
        n.get("name", ""),
        n.get("group", ""),
        n.get("gender", ""),
        ",".join(capable),
        "Y" if n.get("is_night_shift") else "N",
        str(n.get("seniority", 0)),
        juhu_ko,
        "Y" if n.get("juhu_auto_rotate", True) else "N",
        "Y" if n.get("is_trainee") else "N",
        n.get("training_end_date") or "",
        n.get("preceptor_id") or "",
        n.get("start_date") or "",
        n.get("end_date") or "",
    ]


@app.get("/api/nurses/export")
def nurse_export():
    """현재 등록된 간호사 목록을 템플릿 형식 CSV로 내보내기 (호환용 — 기본은 export.xlsx)"""
    import io
    import csv
    from fastapi.responses import Response

    buf = io.StringIO()
    buf.write("\ufeff")  # UTF-8 BOM
    writer = csv.writer(buf)
    writer.writerow(_NURSE_CSV_HEADER)
    for n in db.get_nurses():
        writer.writerow(_nurse_to_row(n))

    content = buf.getvalue().encode("utf-8")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="nurses_current.csv"'
        },
    )


_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_nurse_xlsx(data_rows: list) -> bytes:
    """간호사 명부 xlsx 생성 — 헤더 서식·열 폭·드롭다운(성별/Y·N/요일) + '작성 방법' 시트.
    템플릿(예시 행)과 내보내기(현재 명부)가 같은 형식을 공유해 왕복 편집이 된다."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "간호사 명부"
    ws.append(_NURSE_CSV_HEADER)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4A6FA5")
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for row in data_rows:
        ws.append(list(row))
    widths = [12, 12, 7, 10, 24, 9, 10, 9, 13, 9, 15, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    def add_dropdown(col_name, options):
        dv = DataValidation(type="list", formula1=f'"{options}"',
                            allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv)
        col = get_column_letter(_NURSE_CSV_HEADER.index(col_name) + 1)
        dv.add(f"{col}2:{col}400")

    add_dropdown("성별", "female,male,여,남")
    add_dropdown("야간전담", "Y,N")
    add_dropdown("주휴요일", "일,월,화,수,목,금,토")
    add_dropdown("주휴로테이션", "Y,N")
    add_dropdown("트레이닝", "Y,N")

    guide = wb.create_sheet("작성 방법")
    for line in _NURSE_CSV_GUIDE:
        guide.append([line[0].lstrip("#").strip()])
    guide.column_dimensions["A"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/nurses/template.xlsx")
def nurse_template_xlsx():
    """간호사 일괄 등록용 엑셀 템플릿 — 드롭다운·안내 시트 포함 ('CSV로 저장' 단계 불필요)"""
    from fastapi.responses import Response

    return Response(
        content=_build_nurse_xlsx(_NURSE_CSV_EXAMPLE),
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="nurses_template.xlsx"'},
    )


@app.get("/api/nurses/export.xlsx")
def nurse_export_xlsx():
    """현재 간호사 목록을 템플릿과 같은 형식의 엑셀로 내보내기"""
    from fastapi.responses import Response

    rows = [_nurse_to_row(n) for n in db.get_nurses()]
    return Response(
        content=_build_nurse_xlsx(rows),
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="nurses_current.xlsx"'},
    )


# ── CSV 파싱 헬퍼 (preview / import 양쪽에서 사용) ─────────────────────

_CSV_ENCODINGS = ("utf-8-sig", "cp949", "euc-kr", "utf-8")


def _pick_nurse_sheet_rows(sheets: list) -> list:
    """명부가 든 시트 선택: '이름' 헤더가 있는 첫 시트 → 없으면 데이터 최다 시트."""
    for s in sheets:
        for row in s["rows"][:5]:
            if any(str(c).strip() == "이름" for c in row):
                return s["rows"]
    if not sheets:
        raise HTTPException(400, "파일에서 표를 찾지 못했습니다.")
    return max(
        sheets,
        key=lambda s: sum(1 for r in s["rows"] for c in r if str(c).strip()),
    )["rows"]


def _decode_nurse_table_rows(body: dict) -> list:
    """가져오기 입력(body) → 2D 문자열 rows.
    csv_b64가 xlsx면 매직 바이트(PK)로 감지해 시트 파싱 — 파일명·확장자 무관.
    텍스트면 인코딩 자동 감지(UTF-8/CP949) 후 CSV 파싱."""
    import base64
    import csv as _csv
    import io

    csv_b64 = body.get("csv_b64")
    if csv_b64:
        try:
            raw = base64.b64decode(csv_b64)
        except Exception as e:
            raise HTTPException(400, f"base64 디코딩 실패: {e}")
        if raw[:4] == b"PK\x03\x04":  # zip 컨테이너 = xlsx/xlsm
            return _pick_nurse_sheet_rows(_parse_xlsx_sheets(raw))
        text = None
        for enc in _CSV_ENCODINGS:
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(400, "지원되지 않는 인코딩 (UTF-8 / CP949 모두 실패)")
    else:
        text = body.get("csv", "")
        if not text:
            raise HTTPException(400, "CSV 내용이 비어 있습니다.")

    if text.startswith("\ufeff"):
        text = text[1:]
    return [r for r in _csv.reader(io.StringIO(text))]


def _parse_nurses_rows(rows_in: list) -> dict:
    """2D 문자열 rows → {nurses, errors[{row,col,got,expected,message}]}.
    id 비어 있으면 자동 생성. '#' 시작 행(안내 주석)·빈 행은 무시.
    CSV·xlsx 어느 쪽 입력이든 _decode_nurse_table_rows가 rows로 맞춰 준다."""
    import uuid

    rows = []
    for r in rows_in:
        cells = ["" if c is None else str(c) for c in r]
        if cells and cells[0].lstrip().startswith("#"):
            continue
        if not any(c.strip() for c in cells):
            continue
        rows.append(cells)

    if len(rows) < 2:
        return {
            "nurses": [],
            "errors": [{
                "row": 0, "col": "_file", "got": "",
                "expected": "헤더 행 + 최소 1명",
                "message": "헤더 행과 최소 1명의 데이터 행이 필요합니다.",
            }],
        }

    header = [c.strip() for c in rows[0]]
    data_rows = rows[1:]

    if "이름" not in header:
        return {
            "nurses": [],
            "errors": [{
                "row": 1, "col": "이름", "got": ",".join(header)[:120],
                "expected": "헤더에 '이름' 포함",
                "message": (
                    "필수 컬럼 '이름'이 헤더에 없습니다. "
                    "엑셀이 다른 인코딩으로 저장한 경우, 템플릿(📥)을 다시 받아 "
                    "수정 후 업로드해 주세요."
                ),
            }],
        }

    juhu_map = {"일": 0, "월": 1, "화": 2, "수": 3, "목": 4, "금": 5, "토": 6}

    def col(row, name, default=""):
        if name in header:
            try:
                return row[header.index(name)].strip()
            except IndexError:
                return default
        return default

    def yn(val, default=False):
        v = (val or "").strip().upper()
        if v in ("Y", "YES", "TRUE", "1", "O"):
            return True
        if v in ("N", "NO", "FALSE", "0", "X"):
            return False
        return default

    def err(row_idx, col_name, got, expected, msg):
        return {"row": row_idx, "col": col_name, "got": got,
                "expected": expected, "message": msg}

    nurses = []
    errors = []

    for idx, row in enumerate(data_rows, start=2):
        name = col(row, "이름")
        if not name:
            errors.append(err(idx, "이름", "", "비어있지 않음",
                              f"{idx}행: 이름이 비어 있어 건너뜀"))
            continue

        nid = col(row, "id")
        if not nid:
            nid = f"n_{uuid.uuid4().hex[:10]}"

        capable_str = col(row, "가능근무", "DC,D,EC,E,NC,N")
        capable = [s.strip() for s in capable_str.split(",") if s.strip()]

        juhu_ko = col(row, "주휴요일")
        juhu_day = juhu_map.get(juhu_ko) if juhu_ko else None
        if juhu_ko and juhu_day is None:
            errors.append(err(idx, "주휴요일", juhu_ko,
                              "일/월/화/수/목/금/토 또는 빈칸",
                              f"{idx}행: 주휴요일 '{juhu_ko}' 인식 불가 → 임의로 처리"))

        seniority_str = col(row, "시니어리티", "0")
        try:
            seniority = int(seniority_str) if seniority_str else 0
        except ValueError:
            seniority = 0
            errors.append(err(idx, "시니어리티", seniority_str, "정수",
                              f"{idx}행: 시니어리티 정수 변환 실패 → 0 처리"))

        def chk_date(field):
            v = col(row, field)
            if not v:
                return None
            # 자릿수만 검사하면 2026-02-31 같은 존재하지 않는 날짜가 통과해
            # 전입/전출·트레이닝 종료 제약이 조용히 무시된다 — 실제 달력 검증
            try:
                datetime.strptime(v, "%Y-%m-%d")
            except ValueError:
                errors.append(err(idx, field, v, "YYYY-MM-DD",
                                  f"{idx}행: {field} 형식/날짜 오류 → 빈값 처리"))
                return None
            return v

        gender_raw = col(row, "성별", "female").lower()
        if gender_raw == "여":
            gender_raw = "female"
        elif gender_raw == "남":
            gender_raw = "male"
        elif gender_raw == "":
            gender_raw = "female"
        elif gender_raw not in ("female", "male"):
            errors.append(err(idx, "성별", gender_raw, "female/male/여/남",
                              f"{idx}행: 성별 '{gender_raw}' → female 처리"))
            gender_raw = "female"

        nurses.append({
            "id": nid,
            "name": name,
            "group": col(row, "그룹"),
            "gender": gender_raw,
            "capable_shifts": capable,
            "is_night_shift": yn(col(row, "야간전담"), False),
            "seniority": seniority,
            "wishes": {},
            "juhu_day": juhu_day,
            "juhu_auto_rotate": yn(col(row, "주휴로테이션"), True),
            "night_months": {},
            "is_trainee": yn(col(row, "트레이닝"), False),
            "training_end_date": chk_date("트레이닝종료일"),
            "preceptor_id": col(row, "프리셉터ID") or None,
            "start_date": chk_date("전입일"),
            "end_date": chk_date("전출일"),
        })

    return {"nurses": nurses, "errors": errors}


_NURSE_DIFF_FIELDS = (
    "name", "group", "gender", "capable_shifts", "is_night_shift",
    "seniority", "juhu_day", "juhu_auto_rotate", "is_trainee",
    "training_end_date", "preceptor_id", "start_date", "end_date",
)


def _resolve_nurse_ids(parsed_nurses: list, existing_nurses: list):
    """파싱한 nurse의 id를 기존 DB와 매칭. id 우선, name+group fallback.
    매칭된 기존 ID 집합 반환."""
    by_id = {n["id"]: n for n in existing_nurses}
    by_name_grp = {(n.get("name", "").strip(), n.get("group", "").strip()): n
                   for n in existing_nurses}

    matched_ids = set()
    for p in parsed_nurses:
        match = None
        if p["id"] in by_id:
            match = by_id[p["id"]]
        else:
            key = (p.get("name", "").strip(), p.get("group", "").strip())
            if key[0] and key in by_name_grp:
                match = by_name_grp[key]
        if match:
            p["id"] = match["id"]
            matched_ids.add(match["id"])
    return matched_ids


def _build_diff(parsed_nurses: list, existing_nurses: list, matched_ids: set,
                replace_all: bool) -> dict:
    by_id = {n["id"]: n for n in existing_nurses}

    will_add = []
    will_update = []
    unchanged = 0

    for p in parsed_nurses:
        old = by_id.get(p["id"])
        if old is None:
            will_add.append({
                "id": p["id"], "name": p["name"], "group": p.get("group", ""),
            })
            continue

        diff_fields = []
        for key in _NURSE_DIFF_FIELDS:
            old_v = old.get(key)
            new_v = p.get(key)
            if isinstance(old_v, list):
                old_v = sorted(old_v or [])
            if isinstance(new_v, list):
                new_v = sorted(new_v or [])
            if old_v != new_v:
                diff_fields.append({
                    "field": key, "old": old.get(key), "new": p.get(key),
                })
        if diff_fields:
            will_update.append({
                "id": p["id"], "name": p["name"],
                "group": p.get("group", ""), "fields": diff_fields,
            })
        else:
            unchanged += 1

    will_delete = []
    if replace_all:
        for n in existing_nurses:
            if n["id"] not in matched_ids:
                will_delete.append({
                    "id": n["id"], "name": n["name"],
                    "group": n.get("group", ""),
                })

    return {
        "will_add": will_add,
        "will_update": will_update,
        "will_delete": will_delete,
        "unchanged_count": unchanged,
    }


@app.post("/api/nurses/import/preview")
def nurse_import_preview(body: dict):
    """업로드 전 미리보기: 어떤 행이 추가/수정/삭제될지 계산만 하고 DB는 건드리지 않음."""
    rows = _decode_nurse_table_rows(body)
    replace_all = bool(body.get("replace_all", False))

    parsed = _parse_nurses_rows(rows)
    existing = db.get_nurses()
    matched_ids = _resolve_nurse_ids(parsed["nurses"], existing)
    diff = _build_diff(parsed["nurses"], existing, matched_ids, replace_all)

    return {
        "ok": True,
        "errors": parsed["errors"],
        "parsed_count": len(parsed["nurses"]),
        "replace_all": replace_all,
        **diff,
    }


@app.post("/api/nurses/import")
def nurse_import(body: dict):
    """
    CSV 파싱 + 매칭 + 저장.
    body: {"csv": "<text>"} 또는 {"csv_b64": "<base64>"} + "replace_all"(bool)
    """
    rows = _decode_nurse_table_rows(body)
    replace_all = bool(body.get("replace_all", False))

    parsed = _parse_nurses_rows(rows)
    nurses_to_save = parsed["nurses"]
    errors = parsed["errors"]

    if not nurses_to_save:
        raise HTTPException(
            400,
            f"유효한 행이 없습니다. 오류 {len(errors)}건: "
            + "; ".join(e["message"] for e in errors[:5]),
        )

    existing = db.get_nurses()
    matched_ids = _resolve_nurse_ids(nurses_to_save, existing)

    try:
        # 업서트를 먼저, 삭제는 마지막에 — 중간 실패 시 '삭제만 반영되고 신규는
        # 미반영'되는 최악의 경우(데이터 소실)를 피한다.
        for nurse in nurses_to_save:
            db.upsert_nurse(nurse)
        if replace_all:
            for n in existing:
                if n["id"] not in matched_ids:
                    db.delete_nurse(n["id"])
    except Exception as e:
        raise HTTPException(500, f"DB 저장 실패: {e}")

    return {
        "ok": True,
        "imported": len(nurses_to_save),
        "errors": errors,
        "replaced": replace_all,
    }


# ── 표 파일 파싱 (붙여넣기 모달의 '파일에서 읽기') ────────────────────────────

_TABLE_FILE_MAX_CELLS = 200_000  # 시트당 셀 상한 — 실수로 올린 초대형 파일 방어


def _cell_to_str(v) -> str:
    """엑셀 셀 값 → 문자열. 날짜는 ISO(YYYY-MM-DD)로 — 클라이언트 날짜 해석이 월·연도까지 받게."""
    import datetime as _dt

    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.date().isoformat()
        return f"{v.date().isoformat()} {v.hour:02d}:{v.minute:02d}"
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, _dt.time):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _parse_xlsx_sheets(raw: bytes) -> list:
    """xlsx/xlsm 바이트 → [{name, rows}]. 병합 셀은 좌상단 값을 범위 전체로 전파."""
    import io

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl이 설치되어 있지 않습니다 (pip install openpyxl)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 파일을 열 수 없습니다: {e}")

    sheets = []
    for ws in wb.worksheets:
        if ws.max_row * ws.max_column > _TABLE_FILE_MAX_CELLS:
            raise HTTPException(400, f"시트 '{ws.title}'가 너무 큽니다 (셀 {_TABLE_FILE_MAX_CELLS:,}개 초과)")
        grid = [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        # 병합 셀 전파 — 병동 양식은 날짜/이름 헤더가 병합돼 있는 경우가 많다
        for rng in ws.merged_cells.ranges:
            top = _cell_to_str(ws.cell(rng.min_row, rng.min_col).value)
            if not top:
                continue
            for r in range(rng.min_row - 1, rng.max_row):
                for c in range(rng.min_col - 1, rng.max_col):
                    if r < len(grid) and c < len(grid[r]) and not grid[r][c]:
                        grid[r][c] = top
        # 끝쪽 빈 행/열 정리
        while grid and not any(s.strip() for s in grid[-1]):
            grid.pop()
        for row in grid:
            while row and not row[-1].strip():
                row.pop()
        if grid:
            sheets.append({"name": ws.title, "rows": grid})
    return sheets


def _parse_delimited_sheets(raw: bytes, filename: str) -> list:
    """csv/tsv/txt 바이트 → [{name, rows}]. 인코딩 자동 감지(CP949 폴백)."""
    import csv as _csv
    import io

    text = None
    for enc in _CSV_ENCODINGS:
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(400, "지원되지 않는 인코딩 (UTF-8 / CP949 모두 실패)")
    if text.startswith("\ufeff"):
        text = text[1:]

    lines = text.splitlines()
    first_line = lines[0] if lines else ""
    delimiter = "\t" if "\t" in first_line else ","
    rows = [row for row in _csv.reader(io.StringIO(text), delimiter=delimiter)]
    while rows and not any(s.strip() for s in rows[-1]):
        rows.pop()
    return [{"name": filename, "rows": rows}] if rows else []


@app.post("/api/parse-table-file")
def parse_table_file(body: dict):
    """
    xlsx/xlsm/csv/tsv 파일(base64) → 시트별 2D 문자열 그리드.
    붙여넣기 모달의 '파일에서 읽기' — 결과는 클라이언트 매칭 파이프라인에 그대로 공급된다.
    body: {"file_b64": "<base64>", "filename": "표.xlsx"}
    """
    import base64

    filename = str(body.get("filename") or "")
    file_b64 = body.get("file_b64")
    if not file_b64:
        raise HTTPException(400, "file_b64가 비어 있습니다.")
    try:
        raw = base64.b64decode(file_b64)
    except Exception as e:
        raise HTTPException(400, f"base64 디코딩 실패: {e}")

    lower = filename.lower()
    if lower.endswith(".xls"):
        raise HTTPException(400, "구형 xls(97-2003)는 지원되지 않습니다 — 엑셀에서 xlsx로 다시 저장하거나, 표를 복사해 붙여넣으세요.")
    if lower.endswith((".xlsx", ".xlsm")):
        sheets = _parse_xlsx_sheets(raw)
    elif lower.endswith((".csv", ".tsv", ".txt")):
        sheets = _parse_delimited_sheets(raw, filename)
    else:
        raise HTTPException(400, "지원 형식: xlsx, xlsm, csv, tsv, txt")

    return {"ok": True, "sheets": sheets}


@app.get("/health")
def health():
    return {"status": "healthy", "message": "서버가 정상 동작 중입니다."}


# ── 프론트엔드 서빙 ─────────────────────────────────────────────────────────

@app.get("/")
def index():
    html_path = Path(__file__).parent.parent / "frontend" / "index.html"
    if html_path.exists():
        return FileResponse(str(html_path))
    raise HTTPException(status_code=404, detail="index.html not found")


# ── 간호사 API ────────────────────────────────────────────────────────────────

@app.get("/api/nurses")
def get_nurses():
    return db.get_nurses()


@app.post("/api/nurses")
def upsert_nurse(nurse: Nurse):
    db.upsert_nurse(nurse.model_dump())
    return {"ok": True}


@app.delete("/api/nurses/{nurse_id}")
def delete_nurse(nurse_id: str):
    db.delete_nurse(nurse_id)
    return {"ok": True}


@app.post("/api/nurses/reorder")
def reorder_nurses(body: dict):
    ids = body.get("ids", [])
    db.reorder_nurses(ids)
    return {"ok": True}


# ── 규칙 API ──────────────────────────────────────────────────────────────────

@app.get("/api/rules")
def get_rules():
    rules = db.get_rules()
    if not rules:
        # 기본값 반환
        from .models import Rules
        return Rules().model_dump()
    return rules


@app.post("/api/rules")
def save_rules(rules: dict):
    db.save_rules(rules)
    return {"ok": True}


# ── 요구사항 API ──────────────────────────────────────────────────────────────

@app.get("/api/requirements")
def get_requirements():
    req = db.get_requirements()
    if not req:
        from .models import Requirements
        return Requirements().model_dump()
    return req


@app.post("/api/requirements")
def save_requirements(body: dict):
    db.save_requirements(body)
    return {"ok": True}


# ── 근무 API ──────────────────────────────────────────────────────────────────

@app.get("/api/shifts")
def get_shifts():
    return db.list_shifts()


@app.post("/api/shifts")
def save_shift(body: dict):
    db.save_shift(
        code=body["code"],
        name=body["name"],
        period=body["period"],
        is_charge=body.get("is_charge", False),
        hours=body.get("hours", ""),
        color_bg=body.get("color_bg", "#f3f4f6"),
        color_text=body.get("color_text", "#374151"),
        sort_order=body.get("sort_order", 0),
        auto_assign=body.get("auto_assign", True),
    )
    return {"ok": True}


@app.delete("/api/shifts/{code}")
def delete_shift(code: str):
    db.delete_shift(code)
    return {"ok": True}


# ── 스케줄 생성 API ───────────────────────────────────────────────────────────

# 일별 인원 사전 검증
from datetime import date as _date, timedelta as _td

def _attach_wish_boosts(request: GenerateRequest) -> Optional[dict]:
    """직전 달들의 위시 거절 이력에서 간호사별 가중 배수를 산출해 request에 주입.
    배수 = 1 + 0.5×누적 미반영 (최대 3배). '지난달 거절당한 사람 우선'의 자동화."""
    try:
        if not any((n.wishes or {}) for n in request.nurses):
            return None
        ledger = db.compute_wish_ledger(request.year, request.month)
        boosts = {}
        for n in request.nurses:
            ent = ledger.get(n.id)
            if ent:
                rejected = max(0, ent["requested"] - ent["granted"])
                if rejected:
                    boosts[n.id] = round(min(3.0, 1.0 + 0.5 * rejected), 2)
        if boosts:
            request.wish_boosts = boosts
        return {"ledger": ledger, "boosts": boosts}
    except Exception:
        return None


def _attach_relax_boosts(request: GenerateRequest) -> Optional[dict]:
    """직전 달들에 원티드(사전입력)가 완화로 뒤집힌 간호사에게 유지 보너스 배수를 주입.
    배수 = 1 + 0.5×누적 뒤집힘 (최대 3배) — 위시 보정과 같은 공식. '지난달 당한 사람은
    이번 달 더 지킨다'의 자동화 (제1원칙 8, M6 P2①)."""
    try:
        if not request.prev_schedule:
            return None
        ledger = db.compute_relax_ledger(request.year, request.month)
        boosts = {}
        for n in request.nurses:
            ent = ledger.get(n.id)
            if ent and ent.get("overridden"):
                boosts[n.id] = round(min(3.0, 1.0 + 0.5 * ent["overridden"]), 2)
        if boosts:
            request.relax_boosts = boosts
        return {"ledger": ledger, "boosts": boosts}
    except Exception:
        return None


def _attach_fairness_offsets(request: GenerateRequest) -> Optional[dict]:
    """공정성 원장(직전 3개월 저장본 누적)을 켜진 규칙별 오프셋으로 request에 주입.
    night_fairness → fairness_offsets(누적 야간), weekend_fairness → weekend_offsets
    (누적 주말·공휴일 근무일, 합집합). 규칙이 없으면 그 오프셋은 건드리지 않는다.
    저장본 파생이라 별도 기록이 필요 없다 (M6 P3②, 결정 1-21)."""
    try:
        rts = {getattr(r, "rule_type", "") for r in request.scoring_rules
               if getattr(r, "enabled", True)}
        want_n = "night_fairness" in rts
        want_w = "weekend_fairness" in rts
        if not (want_n or want_w):
            return None
        fl = db.compute_fairness_ledger(request.year, request.month)
        out = {"ledger": fl}
        if want_n:
            offsets = {nid: int(ent.get("nights") or 0)
                       for nid, ent in fl.items() if ent.get("nights")}
            if offsets:
                request.fairness_offsets = offsets
                out["fairness_offsets"] = offsets
        if want_w:
            offsets = {nid: int(ent.get("weekend_holiday") or 0)
                       for nid, ent in fl.items() if ent.get("weekend_holiday")}
            if offsets:
                request.weekend_offsets = offsets
                out["weekend_offsets"] = offsets
        return out
    except Exception:
        return None


def _build_wish_report(request: GenerateRequest, result: dict) -> Optional[dict]:
    """생성 결과 대비 위시 반영 리포트 — 간호사별 신청/반영/미반영 목록."""
    if not result.get("success"):
        return None
    sched = result.get("schedule") or {}
    per_nurse = []
    total_req = total_ok = 0
    month_prefix = f"{request.year:04d}-{request.month:02d}-"
    for n in request.nurses:
        wishes = n.wishes or {}
        if not wishes:
            continue
        ns = sched.get(n.id) or {}
        requested = granted = 0
        unmet = []
        leave_filled = []   # OFF 위시를 OF가 아니라 연차·휴가(V/생/…)로 채운 날 — 반영은 맞지만 연차가 소진된다
        for day_str, wish in wishes.items():
            ds = str(day_str)
            if "-" in ds:
                dk = ds
                if not dk.startswith(month_prefix):
                    continue
            else:
                try:
                    dk = f"{month_prefix}{int(ds):02d}"
                except (ValueError, TypeError):
                    continue
            assigned = ns.get(dk, "")
            if not assigned:
                continue
            requested += 1
            if wish == "OFF":
                ok = assigned in db.WISH_REST_LIKE or assigned in db.WISH_LEAVE_LIKE
                if ok and assigned in db.WISH_LEAVE_LIKE:
                    leave_filled.append(dk)
            else:
                ok = (assigned == wish)
            if ok:
                granted += 1
            else:
                unmet.append({"date": dk, "wish": wish, "assigned": assigned})
        if requested:
            total_req += requested
            total_ok += granted
            per_nurse.append({
                "nurse_id": n.id, "name": n.name,
                "requested": requested, "granted": granted, "unmet": unmet,
                "leave_filled": len(leave_filled), "leave_filled_dates": leave_filled,
                "boost": (request.wish_boosts or {}).get(n.id),
            })
    if not per_nurse:
        return None
    per_nurse.sort(key=lambda r: (r["granted"] / r["requested"], -len(r["unmet"])))
    return {"total_requested": total_req, "total_granted": total_ok,
            "total_leave_filled": sum(r["leave_filled"] for r in per_nurse),
            "per_nurse": per_nurse}


def _validate_locked_conflicts(request: GenerateRequest) -> Optional[str]:
    """잠긴 셀의 사전입력이 규칙에 의해 무효화(드롭)되는 충돌 검출.
    '잠금은 완화에서도 고정'이 약속이지만, 공휴일 OF 금지 같은 규칙이 사전입력
    자체를 드롭하면 잠금이 적용될 수 없다 — 조용히 무시하지 말고 경고한다."""
    locked = request.locked_cells or {}
    prev = request.prev_schedule or {}
    holidays = set(request.holidays or [])
    name_by_id = {n.id: n.name for n in request.nurses}
    bad = []
    for nid, cells in locked.items():
        for dt_str, flag in (cells or {}).items():
            if not flag:
                continue
            pre = (prev.get(nid) or {}).get(dt_str)
            if pre == "OF" and dt_str in holidays:
                bad.append(f"  · {name_by_id.get(nid, nid)} {dt_str}: 공휴일 OF 잠금")
    if not bad:
        return None
    return ("⚠ 공휴일에는 OF를 배정할 수 없어 아래 잠금이 적용되지 않습니다 "
            "(법정공휴일 '법' 코드 사용 권장):\n" + "\n".join(bad[:8]))


def _validate_staffing(request: GenerateRequest, leave_shifts: list, rest_shifts: list) -> Optional[str]:
    """
    prev_schedule의 고정 근무를 반영한 후, 각 날짜별 근무 가능 인원이
    요구사항을 충족할 수 있는지 사전 검증.
    부족한 날이 있으면 경고 메시지 반환, 없으면 None.
    """
    weekday_keys = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
    year, month = request.year, request.month
    first = _date(year, month, 1)
    if month == 12:
        last = _date(year + 1, 1, 1) - _td(days=1)
    else:
        last = _date(year, month + 1, 1) - _td(days=1)

    req_dict = request.requirements.model_dump()
    off_shifts = leave_shifts + rest_shifts
    prev = request.prev_schedule or {}
    warnings = []

    cur = first
    while cur <= last:
        dt_str = cur.strftime("%Y-%m-%d")
        weekday_key = weekday_keys[cur.weekday()]
        per_day = request.per_day_requirements or {}
        day_req = per_day.get(dt_str) or req_dict.get(weekday_key, {})

        unavailable = sum(
            1 for nurse in request.nurses
            if prev.get(nurse.id, {}).get(dt_str, "") in off_shifts
        )
        available = len(request.nurses) - unavailable

        total_needed = sum(day_req.get(p, 0) for p in ["D", "E", "N"])
        if available < total_needed:
            warnings.append(
                f"{cur.strftime('%m/%d')}({['월','화','수','목','금','토','일'][cur.weekday()]}): "
                f"필요 {total_needed}명, 가용 {available}명 (부족 {total_needed - available}명)"
            )
        cur += _td(days=1)

    if warnings:
        return "경고: 일부 날짜에 인원이 부족할 수 있습니다.\n" + "\n".join(warnings[:5]) + (
            f"\n... 외 {len(warnings)-5}일" if len(warnings) > 5 else ""
        )
    return None


@app.post("/api/estimate")
def estimate(request: GenerateRequest):
    """스케줄 생성 전 예상 소요시간(초) 반환."""
    try:
        if not request.scoring_rules:
            raw = db.list_scoring_rules()
            request.scoring_rules = [ScoringRule(**r) for r in raw]
        if not request.shifts:
            from .models import ShiftDef
            raw = db.list_shifts()
            request.shifts = [ShiftDef(**s) for s in raw]
        scheduler = NurseScheduler(request)
        return {"estimated_seconds": scheduler.estimate_seconds()}
    except Exception as e:
        logger.error("Server error: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail="서버 오류가 발생했습니다. 다시 시도해주세요.")


@app.post("/api/generate/stop")
def stop_generate():
    """진행 중인 스케줄 생성을 중지하고 지금까지 찾은 최선의 해를 반환하도록 신호."""
    # is_active(어댑터 등록) 대신 is_running(수명주기 래치) 기준 — 모델 빌드
    # 구간(어댑터 등록 전)에 눌린 중지도 플래그로 보존되어, register() 시점에
    # 자동으로 cancel이 전달된다.
    if solver_progress.is_running():
        solver_progress.request_cancel()
        return {"ok": True, "message": "중지 신호를 전송했습니다."}
    return {"ok": False, "message": "진행 중인 생성이 없습니다."}


@app.get("/api/generate/stream")
def generate_stream():
    """SSE: 솔버 로그 + 진행 상황 실시간 스트리밍"""
    def event_gen():
        import time
        # 솔버가 아직 시작되지 않았을 수 있으므로 최대 30초 대기
        waited = 0
        while not solver_progress.is_active() and _log_queue.empty() and waited < 30:
            if solver_progress.is_running():
                # generate()가 호출됨 → 솔버 시작 대기
                time.sleep(0.2)
                waited += 0.2
            else:
                # generate() 자체가 아직 호출 안 됨 → 짧게 대기
                time.sleep(0.5)
                waited += 0.5
        while True:
            # 로그 메시지 우선 드레인
            try:
                item = _log_queue.get(timeout=0.05)
                yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"
                continue
            except queue.Empty:
                pass
            # 솔버 종료 + 큐 비어있으면 done
            if not solver_progress.is_active() and _log_queue.empty() and not solver_progress.is_running():
                yield "data: {\"type\":\"done\"}\n\n"
                break
            # 1초 heartbeat — 현재 progress 전송
            prog = solver_progress.get_progress()
            prog["type"] = "progress"
            yield f"data: {json.dumps(prog)}\n\n"
            time.sleep(1)
    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/generate/progress")
def get_generate_progress():
    """생성 중 실시간 진행 상황 반환 (2초 간격 폴링용)"""
    return solver_progress.get_progress()


def _run_race(request: GenerateRequest) -> dict:
    """두 엔진 레이스 — HiGHS와 CP-SAT를 동시 실행, 먼저 성공한 쪽을 채택하고
    패자는 즉시 중지(cancel_all_adapters, 사용자 취소 플래그는 건드리지 않음).

    안전성: 승자 확정 후 양 스레드를 bounded join으로 정리한다. 드물게 패자(특히
    incumbent 없는 CP-SAT)가 콜백 게이트 때문에 즉시 멈추지 못하면 데몬 스레드로
    남지만, generate()의 solver_progress.end()가 레지스트리를 비우므로 다음 생성이
    409로 막히거나 상태가 오염되지 않는다(패자 결과는 폐기)."""
    from .scheduler_cpsat import CpSatScheduler
    results: dict = {}
    first_success = threading.Event()
    decide_lock = threading.Lock()
    winner = {"name": None}

    def _run(name, factory):
        try:
            res = factory(request).solve()
        except Exception as e:  # 한 엔진이 터져도 다른 엔진 결과로 진행
            res = {"success": False, "message": f"[{name}] 오류: {e}", "schedule": {}}
        results[name] = res
        if res.get("success"):
            with decide_lock:
                if winner["name"] is None:
                    winner["name"] = name
                    first_success.set()
                    solver_progress.cancel_all_adapters()  # 패자 중지

    threads = []
    for name, factory in (("HiGHS", NurseScheduler), ("CP-SAT", CpSatScheduler)):
        t = threading.Thread(target=_run, args=(name, factory), daemon=True, name=f"race-{name}")
        t.start()
        threads.append(t)

    # 첫 성공 OR 둘 다 종료까지 대기 — 솔버 자체 time_limit가 전체 백스톱.
    # (둘 다 빠르게 실패하면 first_success가 안 떠도 all-dead로 즉시 빠져나가야 함)
    import time as _time
    overall = float(getattr(request, "time_limit", 1200) or 1200) + 60
    deadline = _time.monotonic() + overall
    while not first_success.is_set() and any(t.is_alive() for t in threads):
        if _time.monotonic() > deadline:
            break
        first_success.wait(timeout=0.1)
    solver_progress.cancel_all_adapters()  # 안전망: 승자 유무와 무관하게 남은 어댑터 취소
    for t in threads:  # 패자 정리(HiGHS는 즉시, CP-SAT는 다음 incumbent까지) — 최대 30s 바운드
        t.join(timeout=30)

    win = winner["name"]
    if win and results.get(win, {}).get("success"):
        res = dict(results[win])
        res["race_winner"] = win
        base = res.get("message", "")
        res["message"] = f"[{win} 우승] {base}" if base else f"{win} 엔진이 먼저 해를 찾았습니다."
        return res
    # 성공한 엔진이 없음 → 더 유용한 메시지(둘 다 infeasible/timeout)를 HiGHS 우선 반환
    for name in ("HiGHS", "CP-SAT"):
        if name in results:
            res = dict(results[name])
            res["race_winner"] = None
            return res
    return {"success": False, "message": "레이스 실패 (두 엔진 모두 결과 없음).", "schedule": {}}


@app.post("/api/generate")
def generate(request: GenerateRequest):
    global _last_mip_gap, _last_generate_result
    # 이전 솔버가 아직 돌고 있으면 거부 (체크-시작을 원자적으로)
    if not solver_progress.try_begin():
        raise HTTPException(status_code=409, detail="이미 생성이 진행 중입니다. 중지 후 다시 시도하세요.")
    _last_mip_gap = None
    _last_generate_result = None  # 새 생성 시작 시 이전 결과 초기화
    try:
        # shifts가 비어있으면 DB에서 로드
        if not request.scoring_rules:
            raw = db.list_scoring_rules()
            request.scoring_rules = [ScoringRule(**r) for r in raw]

        # shifts가 비어있으면 DB에서 로드
        if not request.shifts:
            from .models import ShiftDef
            raw = db.list_shifts()
            request.shifts = [ShiftDef(**s) for s in raw]

        leave_shifts = [s.code for s in request.shifts if s.period == "leave"]
        rest_shifts  = [s.code for s in request.shifts if s.period == "rest"]

        warning = _validate_staffing(request, leave_shifts, rest_shifts)
        lock_warn = _validate_locked_conflicts(request)
        if lock_warn:
            warning = (warning + "\n\n" + lock_warn) if warning else lock_warn
        wish_ctx = _attach_wish_boosts(request)
        relax_ctx = _attach_relax_boosts(request)
        # 공정성 원장 — night_fairness·weekend_fairness가 (직전 3개월 누적 + 당월)
        # 편차를 줄이도록 누적 오프셋을 자동 주입 (저장본 파생 — 별도 기록 불필요)
        fair_ctx = _attach_fairness_offsets(request)
        if request.solver == "race":
            result = _run_race(request)
        elif request.solver == "cpsat":
            from .scheduler_cpsat import CpSatScheduler
            result = CpSatScheduler(request).solve()
        else:
            result = NurseScheduler(request).solve()

        # MIP gap 및 중지 여부 추가 (race는 _run_race가 승자 gap을 이미 결과에 담음 →
        # HiGHS 패자의 stale _last_mip_gap로 CP-SAT 승자 gap을 덮어쓰지 않도록 제외)
        if _last_mip_gap is not None and request.solver != "race":
            import math
            if math.isfinite(_last_mip_gap):
                result["mip_gap_percent"] = round(_last_mip_gap * 100, 2)
        if solver_progress.is_cancelled() and result.get("success"):
            result["stopped"] = True

        if warning and not result.get("success"):
            result["message"] = warning + "\n\n" + result.get("message", "")
        elif warning:
            result["warning"] = warning

        # 위시 반영 리포트 — 거절의 투명성 (누가 몇 건 신청·반영됐는지 + 보정 배수)
        try:
            wr = _build_wish_report(request, result)
            if wr:
                if wish_ctx:
                    wr["ledger"] = wish_ctx.get("ledger") or {}
                result["wish_report"] = wr
        except Exception:
            pass
        # 공정성 원장 오프셋 — ⚖ 공정성 카드가 '누적이 이번 생성에 반영됐다'를 보여주도록 노출
        if fair_ctx and result.get("success"):
            for _k in ("fairness_offsets", "weekend_offsets"):
                if fair_ctx.get(_k):
                    result[_k] = fair_ctx[_k]
        # 생성 리포트 — run 레코더 집계 (end() 호출 전에 만들어야 함)
        try:
            report = solver_progress.build_report(result)
            if report:
                if relax_ctx and relax_ctx.get("boosts"):
                    result["relax_boosts"] = relax_ctx["boosts"]
                if result.get("relaxed_cells"):
                    report["relax_attempted"] = True
                try:
                    note, recent = db.insert_generation_run({
                        "year": request.year, "month": request.month,
                        "solver": request.solver,
                        "success": 1 if result.get("success") else 0,
                        "stopped": 1 if result.get("stopped") else 0,
                        "relaxed": 1 if report.get("relax_attempted") else 0,
                        "duration_s": report.get("duration_seconds"),
                        "final_gap": report.get("final_gap_percent"),
                        "num_vars": report.get("num_vars"),
                        "num_constraints": report.get("num_constraints"),
                        "nurse_count": len(request.nurses or []),
                        "pre_filled": sum(len(v) for v in (request.prev_schedule or {}).values()),
                        "time_limit": request.time_limit,
                        "mip_gap": request.mip_gap,
                    })
                    if note:
                        report["history_note"] = note
                    report["recent_runs"] = recent
                except Exception:
                    pass
                result["generation_report"] = report
        except Exception:
            pass
        _last_generate_result = result  # 결과 보관
        solver_progress.end()
        return result
    except Exception as e:
        solver_progress.end()
        _last_generate_result = {"success": False, "message": str(e), "schedule": {}}
        logger.error("Server error: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail="서버 오류가 발생했습니다. 다시 시도해주세요.")


@app.get("/api/generate/result")
def get_generate_result():
    """마지막 생성 결과 조회 (새로고침 복구용)"""
    if solver_progress.is_running():
        return {"status": "running"}
    if _last_generate_result is not None:
        return {"status": "done", "result": _last_generate_result}
    return {"status": "idle"}


@app.post("/api/diagnose")
def diagnose(request: GenerateRequest):
    """CP-SAT assumptions 기반 정밀 충돌 분석.
    어느 엔진으로 생성하다 infeasible이 나든 '어느 제약이 동시 충족 불가인지' 짚는다."""
    from .conflict_analyzer import analyze_conflicts
    if not request.shifts:
        from .models import ShiftDef
        request.shifts = [ShiftDef(**s) for s in db.list_shifts()]
    if not request.scoring_rules:
        request.scoring_rules = [ScoringRule(**r) for r in db.list_scoring_rules()]
    try:
        return analyze_conflicts(request)
    except Exception as e:
        logger.error("diagnose error: %s\n%s", e, traceback.format_exc())
        return {"conflicts": [], "message": f"충돌 분석 중 오류: {e}"}


_feasibility_lock = threading.Lock()  # 프로브 동시 실행 방지 (편집 연타 시 CPU 보호)


@app.post("/api/feasibility")
def feasibility(request: GenerateRequest):
    """사전입력 실시간 신호등 — 하드 제약만 1회 풀이(기본 5초 캡).
    생성 솔버 실행 중이거나 다른 프로브가 도는 중엔 busy 반환 (프론트가 재시도)."""
    if solver_progress.is_running() or not _feasibility_lock.acquire(blocking=False):
        return {"status": "busy"}
    try:
        from .conflict_analyzer import check_feasibility
        if not request.shifts:
            from .models import ShiftDef
            request.shifts = [ShiftDef(**s) for s in db.list_shifts()]
        if not request.scoring_rules:
            request.scoring_rules = [ScoringRule(**r) for r in db.list_scoring_rules()]
        return check_feasibility(request)
    except Exception as e:
        logger.error("feasibility error: %s\n%s", e, traceback.format_exc())
        # 배지는 '미확정'으로 — 실시간 피드백이 500으로 UI를 흔들지 않게
        return {"status": "unknown", "conflicts": []}
    finally:
        _feasibility_lock.release()


@app.post("/api/suggest-fix")
def suggest_fix(request: GenerateRequest):
    """최소 수정 처방(MCS): 어떤 사전입력을 빼거나 인원을 얼마나 늘리면 생성 가능한지."""
    from .conflict_analyzer import suggest_correction
    if not request.shifts:
        from .models import ShiftDef
        request.shifts = [ShiftDef(**s) for s in db.list_shifts()]
    if not request.scoring_rules:
        request.scoring_rules = [ScoringRule(**r) for r in db.list_scoring_rules()]
    try:
        return suggest_correction(request)
    except Exception as e:
        logger.error("suggest-fix error: %s\n%s", e, traceback.format_exc())
        return {"fixable": False, "message": f"수정 처방 계산 중 오류: {e}"}


# ── 스케줄 저장/관리 API ──────────────────────────────────────────────────────

@app.get("/api/schedules")
def list_schedules():
    return db.list_schedules()


@app.post("/api/schedules")
def save_schedule(body: ScheduleSave):
    sid = db.save_schedule(
        year=body.year,
        month=body.month,
        data={
            "nurses": [n.model_dump() for n in body.nurses],
            "requirements": body.requirements.model_dump(),
            "rules": body.rules.model_dump(),
            "schedule": body.schedule,
            "prev_schedule": body.prev_schedule or {},
            "nurse_scores": body.nurse_scores or {},
            "nurse_score_details": body.nurse_score_details or {},
            "locked_cells": body.locked_cells or {},
            "cell_notes": body.cell_notes or {},
            "holidays": body.holidays or [],
            "prev_day_reqs": body.prev_day_reqs or {},
            "prev_month_nights": body.prev_month_nights or {},
            "solver_log": body.solver_log or "",
        },
        name=body.name,
    )
    return {"id": sid}


@app.get("/api/schedules/{schedule_id}")
def load_schedule(schedule_id: int):
    result = db.load_schedule(schedule_id)
    if not result:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다.")
    return result


@app.delete("/api/schedules/{schedule_id}")
def delete_schedule(schedule_id: int):
    db.delete_schedule(schedule_id)
    return {"ok": True}


# ── 사전입력 저장/관리 API ─────────────────────────────────────────────────────

@app.get("/api/prev_schedules")
def list_prev_schedules():
    return db.list_prev_schedules()


@app.post("/api/prev_schedules")
def save_prev_schedule(body: dict):
    # 저장 전 정규화: 삭제된 간호사(유령) 엔트리 제거
    data = body.get("data") or {}
    valid_ids = set(n["id"] for n in db.get_nurses())
    for key in ("schedule", "prev_month_nights", "locked_cells", "cell_notes", "relaxed_cells"):
        sub = data.get(key)
        if isinstance(sub, dict):
            data[key] = {k: v for k, v in sub.items() if k in valid_ids}
    pid = db.save_prev_schedule(
        year=body["year"],
        month=body["month"],
        data=data,
        name=body.get("name"),
    )
    return {"id": pid}


@app.get("/api/prev_schedules/{prev_id}")
def load_prev_schedule(prev_id: int):
    result = db.load_prev_schedule(prev_id)
    if not result:
        raise HTTPException(status_code=404, detail="사전입력을 찾을 수 없습니다.")
    return result


@app.delete("/api/prev_schedules/{prev_id}")
def delete_prev_schedule(prev_id: int):
    db.delete_prev_schedule(prev_id)
    return {"ok": True}


# ── 배점 규칙 API ─────────────────────────────────────────────────────────────

@app.get("/api/scoring_rules")
def get_scoring_rules():
    return db.list_scoring_rules()


@app.post("/api/scoring_rules")
def save_scoring_rule(body: dict):
    rid = db.save_scoring_rule(
        name=body["name"],
        rule_type=body["rule_type"],
        params=body.get("params", {}),
        score=body["score"],
        enabled=body.get("enabled", True),
        sort_order=body.get("sort_order", 0),
        rule_id=body.get("id"),
    )
    return {"id": rid}


@app.delete("/api/scoring_rules/{rule_id}")
def delete_scoring_rule(rule_id: int):
    db.delete_scoring_rule(rule_id)
    return {"ok": True}
